import os
import csv
import frappe


def read_zoho_file(file_path):
    """Try CSV first always, then Excel as fallback."""

    # -- Always try CSV first regardless of extension --
    try:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        if rows and len(rows[0]) > 1:
            frappe.log_error(str(list(rows[0].keys())), "ZOHO COLUMNS DETECTED (CSV)")
            return rows
    except Exception as e:
        frappe.log_error(str(e), "ZOHO CSV ERROR")

    # -- Excel fallback --
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (v if v is not None else "") for i, v in enumerate(row)})
        wb.close()
        frappe.log_error(str(headers), "ZOHO COLUMNS DETECTED (Excel)")
        return rows
    except Exception as e:
        frappe.log_error(str(e), "ZOHO EXCEL ERROR")

    frappe.throw("Cannot parse file - must be CSV or Excel (.xlsx)")


def safe(row, *keys):
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip() not in ("", "None", "nan"):
            return str(v).strip()
    return None


@frappe.whitelist()
def process_zoho_import(docname):
    doc = frappe.get_doc("Zoho Import", docname)
    doc.status = "Processing"
    doc.error_log = ""
    doc.imported_rows = 0
    doc.new_agencies = 0
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    try:
        if not doc.import_file:
            frappe.throw("No file attached.")

        file_doc = frappe.get_doc("File", {"file_url": doc.import_file})
        file_path = file_doc.get_full_path()

        # Log file info for debugging
        frappe.log_error(f"Path: {file_path} | Ext: {os.path.splitext(file_path)[1]}", "ZOHO FILE INFO")

        rows = read_zoho_file(file_path)

        if not rows:
            frappe.throw("File is empty or cannot be read.")

        imported = 0
        new_count = 0
        updated = 0
        skipped = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            agency_name = safe(row, "Full Name", "Agency Name", "Name")
            if not agency_name:
                skipped += 1
                continue

            pcc_code    = safe(row, "PCC Code")
            iata_number = safe(row, "IATA Number")
            email       = safe(row, "Email", "Secondary Email")
            phone       = safe(row, "Phone", "Home Phone", "Other Phone")
            mobile      = safe(row, "Mobile")
            zoho_id     = safe(row, "Record Id")
            created_t   = safe(row, "Created Time")
            modified_t  = safe(row, "Modified Time")

            try:
                exists = frappe.db.get_value("Agency", {"agency_name": agency_name}, "name")
                if not exists:
                    agency_doc = frappe.get_doc({
                        "doctype":            "Agency",
                        "agency_name":        agency_name,
                        "pcc_code":           pcc_code,
                        "iata_number":        iata_number,
                        "email":              email,
                        "phone":              phone,
                        "mobile":             mobile,
                        "zoho_record_id":     zoho_id,
                        "zoho_created_time":  created_t,
                        "zoho_modified_time": modified_t,
                    })
                    agency_doc.insert(ignore_permissions=True)
                    new_count += 1
                else:
                    existing = frappe.get_doc("Agency", exists)
                    changed = False
                    for field, value in [
                        ("pcc_code",           pcc_code),
                        ("iata_number",        iata_number),
                        ("email",              email),
                        ("phone",              phone),
                        ("mobile",             mobile),
                        ("zoho_record_id",     zoho_id),
                        ("zoho_created_time",  created_t),
                        ("zoho_modified_time", modified_t),
                    ]:
                        if value and not getattr(existing, field, None):
                            setattr(existing, field, value)
                            changed = True
                    if changed:
                        existing.save(ignore_permissions=True)
                        updated += 1

                imported += 1

            except Exception as row_err:
                errors.append(f"Row {idx} ({agency_name}): {str(row_err)}")
                frappe.log_error(frappe.get_traceback(), f"ZOHO ROW {idx} ERROR")

        frappe.db.commit()

        summary = f"Completed\\nImported: {imported} | New: {new_count} | Updated: {updated} | Skipped: {skipped}"
        if errors:
            summary += "\\n\\nRow Errors:\\n" + "\\n".join(errors[:20])

        doc = frappe.get_doc("Zoho Import", docname)
        doc.imported_rows = imported
        doc.new_agencies  = new_count
        doc.status        = "Completed"
        doc.error_log     = summary
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "status":   "success",
            "imported": imported,
            "new":      new_count,
            "updated":  updated,
            "skipped":  skipped,
            "errors":   len(errors),
        }

    except Exception:
        doc.status    = "Failed"
        doc.error_log = frappe.get_traceback()
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        raise

