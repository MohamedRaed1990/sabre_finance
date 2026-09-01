import openpyxl
import frappe


def safe(val):
    if val is not None and str(val).strip() not in ("", "None", "nan"):
        return str(val).strip()
    return None


def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(str(val).replace(',', '').strip()))
    except (ValueError, TypeError):
        return None


def find_col_indices(header_row, wanted):
    """
    header_row: tuple of cell values from the header row
    wanted: dict like {"sc_code": "SC Code", "iata_name": "IATA Name", ...}
    returns: dict {"sc_code": 3, "iata_name": 7, ...} (0-based column index) or raises
    """
    normalized = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            normalized[str(cell).strip().lower()] = idx

    result = {}
    missing = []
    for key, header_name in wanted.items():
        idx = normalized.get(header_name.strip().lower())
        if idx is None:
            missing.append(header_name)
        else:
            result[key] = idx

    if missing:
        frappe.throw(f"Could not find the following columns in the file: {', '.join(missing)}")

    return result


def is_overall_row(sc_code, iata_name):
    """Detect the 'Overall' summary row so it is never imported as data."""
    for value in (sc_code, iata_name):
        if value and value.strip().lower() == "overall":
            return True
    return False


@frappe.whitelist()
def process_midt_import(docname):
    doc = frappe.get_doc("MIDT Import", docname)
    doc.status = "Processing"
    doc.error_log = ""
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    try:
        file_url = doc.midt_file
        if not file_url:
            frappe.throw("Please attach a MIDT file first.")

        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_path = file_doc.get_full_path()

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

        rows_iter = ws.iter_rows(values_only=True)

        header_row = None
        for row in rows_iter:
            if row and any(cell is not None and str(cell).strip() != "" for cell in row):
                header_row = row
                break
        if not header_row:
            frappe.throw("File is empty or has no header row.")

        cols = find_col_indices(header_row, {
            "sc_code": "SC Code",
            "iata_name": "IATA Name",
            "sabre_bookings": "Sabre Bookings",
        })

        skipped = 0

        # An sc_code can appear in several different places in the
        # sheet. Aggregate all rows first: keep the first iata_name
        # seen for each sc_code, and sum sabre_bookings across every
        # occurrence, before writing anything to the database.
        grouped = {}

        for row in rows_iter:
            if not row or not any(row):
                continue

            sc_code = safe(row[cols["sc_code"]]) if cols["sc_code"] < len(row) else None
            iata_name = safe(row[cols["iata_name"]]) if cols["iata_name"] < len(row) else None
            bookings = safe_int(row[cols["sabre_bookings"]]) if cols["sabre_bookings"] < len(row) else None

            first_cell = safe(row[0]) if row else None
            if first_cell and "grand total" in first_cell.lower():
                break

            if is_overall_row(sc_code, iata_name):
                skipped += 1
                continue

            if not sc_code:
                skipped += 1
                continue

            g = grouped.setdefault(sc_code, {"iata_name": None, "sabre_bookings": 0})
            if not g["iata_name"] and iata_name:
                g["iata_name"] = iata_name
            g["sabre_bookings"] += bookings or 0

        wb.close()

        imported = 0
        duplicates = 0

        for sc_code, g in grouped.items():
            exists = frappe.db.exists("MIDT Data", {
                "month": doc.month,
                "year": doc.year,
                "sc_code": sc_code,
            })
            if exists:
                duplicates += 1
                continue

            data_doc = frappe.get_doc({
                "doctype": "MIDT Data",
                "midt_import": docname,
                "month": doc.month,
                "year": doc.year,
                "sc_code": sc_code,
                "iata_name": g["iata_name"],
                "sabre_bookings": g["sabre_bookings"],
            })
            data_doc.insert(ignore_permissions=True)
            imported += 1

            if imported % 200 == 0:
                frappe.db.commit()

        doc.status = "Completed"
        doc.error_log = f"Imported: {imported} | Skipped: {skipped} | Duplicates: {duplicates}"
        doc.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "status": "success",
            "imported": imported,
            "skipped": skipped,
            "duplicates": duplicates,
        }

    except Exception:
        doc.status = "Failed"
        doc.error_log = frappe.get_traceback()
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        raise
